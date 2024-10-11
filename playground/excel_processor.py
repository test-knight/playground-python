import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


def process_xlsx(filename):
    new_filename = filename.rstrip('.xlsx') + "_copy.xlsx"
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')
    wb.save(new_filename)


path = Path()
for file in path.glob('*.xlsx'):
    process_xlsx(file.name)
